"""
File content extraction for uploaded documents.

Supported types:
  PDF          → text via pypdf (falls back to OCR service)
  Excel/CSV    → JSON rows via openpyxl / csv
  Word (.docx) → text via python-docx
  Image        → text via OCR service (Tesseract)
  URL          → text via httpx + basic HTML strip
  Plain text   → direct decode
"""

import csv
import io
import json
import os
import re
from typing import Any

import httpx

OCR_URL = os.environ.get("OCR_SVC_URL", "http://ocr-svc:8094")


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def _extract_pdf(data: bytes) -> dict[str, Any]:
    try:
        import pypdf
        reader = pypdf.PdfReader(io.BytesIO(data))
        pages = []
        for i, page in enumerate(reader.pages):
            text = page.extract_text() or ""
            pages.append({"page": i + 1, "text": text.strip()})
        full_text = "\n\n".join(p["text"] for p in pages if p["text"])
        return {"format": "pdf", "pages": len(pages), "text": full_text, "page_texts": pages}
    except ImportError:
        pass

    # Fallback: OCR service
    try:
        resp = httpx.post(
            f"{OCR_URL}/ocr/extract",
            files={"file": ("document.pdf", data, "application/pdf")},
            timeout=60,
        )
        if resp.status_code == 200:
            d = resp.json()
            return {"format": "pdf_ocr", "text": d.get("text", ""), "pages": d.get("pages", 1)}
    except Exception:
        pass

    return {"format": "pdf", "text": "[PDF extraction failed — no pypdf or OCR available]", "pages": 0}


# ---------------------------------------------------------------------------
# Excel / CSV
# ---------------------------------------------------------------------------

def _extract_excel(data: bytes, filename: str = "") -> dict[str, Any]:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext == "csv":
        try:
            text = data.decode("utf-8", errors="replace")
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
            summary = f"CSV with {len(rows)} rows, columns: {', '.join(reader.fieldnames or [])}"
            return {"format": "csv", "rows": rows[:500], "total_rows": len(rows),
                    "columns": reader.fieldnames or [], "text": summary,
                    "json_preview": json.dumps(rows[:20], indent=2)}
        except Exception as e:
            return {"format": "csv", "text": data.decode("utf-8", errors="replace"), "error": str(e)}

    # Excel
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheets = {}
        for name in wb.sheetnames:
            ws = wb[name]
            rows = []
            headers: list[str] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    headers = [str(c) if c is not None else f"col_{j}" for j, c in enumerate(row)]
                else:
                    rows.append(dict(zip(headers, [str(c) if c is not None else "" for c in row])))
                if i >= 1000:
                    break
            sheets[name] = {"headers": headers, "rows": rows, "total_rows": ws.max_row or 0}
        first_sheet = sheets.get(wb.sheetnames[0], {})
        text = (
            f"Excel workbook with {len(wb.sheetnames)} sheet(s): {', '.join(wb.sheetnames)}. "
            f"First sheet '{wb.sheetnames[0]}': {first_sheet.get('total_rows', 0)} rows, "
            f"columns: {', '.join(first_sheet.get('headers', []))}"
        )
        return {
            "format": "excel",
            "sheets": list(wb.sheetnames),
            "data": sheets,
            "text": text,
            "json_preview": json.dumps(first_sheet.get("rows", [])[:20], indent=2),
        }
    except ImportError:
        return {"format": "excel", "text": "[Excel extraction requires openpyxl]"}
    except Exception as e:
        return {"format": "excel", "text": f"[Excel extraction error: {e}]"}


# ---------------------------------------------------------------------------
# Word (.docx)
# ---------------------------------------------------------------------------

def _extract_word(data: bytes) -> dict[str, Any]:
    try:
        import docx
        doc = docx.Document(io.BytesIO(data))
        paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
        text = "\n\n".join(paragraphs)
        return {"format": "docx", "paragraphs": len(paragraphs), "text": text}
    except ImportError:
        return {"format": "docx", "text": "[Word extraction requires python-docx]"}
    except Exception as e:
        return {"format": "docx", "text": f"[Word extraction error: {e}]"}


# ---------------------------------------------------------------------------
# Image (via OCR service)
# ---------------------------------------------------------------------------

def _extract_image(data: bytes, content_type: str, filename: str = "") -> dict[str, Any]:
    try:
        resp = httpx.post(
            f"{OCR_URL}/ocr/extract",
            files={"file": (filename or "image.png", data, content_type)},
            timeout=60,
        )
        if resp.status_code == 200:
            d = resp.json()
            return {"format": "image_ocr", "text": d.get("text", ""), "confidence": d.get("confidence")}
        return {"format": "image_ocr", "text": f"[OCR failed: {resp.status_code}]"}
    except Exception as e:
        return {"format": "image_ocr", "text": f"[OCR service unavailable: {e}]"}


# ---------------------------------------------------------------------------
# URL fetch
# ---------------------------------------------------------------------------

def _extract_url(url: str) -> dict[str, Any]:
    try:
        resp = httpx.get(url, follow_redirects=True, timeout=20,
                         headers={"User-Agent": "AtomPlatform/1.0"})
        ct = resp.headers.get("content-type", "")
        raw = resp.text

        if "text/html" in ct:
            # Basic HTML → text strip
            text = re.sub(r"<style[^>]*>.*?</style>", " ", raw, flags=re.DOTALL | re.IGNORECASE)
            text = re.sub(r"<script[^>]*>.*?</script>", " ", text, flags=re.DOTALL | re.IGNORECASE)
            text = re.sub(r"<[^>]+>", " ", text)
            text = re.sub(r"\s{3,}", "\n\n", text).strip()
            text = text[:8000]  # cap at 8k chars
            return {"format": "html", "url": url, "text": text, "content_type": ct}

        if "json" in ct:
            try:
                parsed = json.loads(raw)
                return {"format": "json", "url": url, "text": json.dumps(parsed, indent=2)[:8000], "data": parsed}
            except Exception:
                pass

        return {"format": "text", "url": url, "text": raw[:8000], "content_type": ct}
    except Exception as e:
        return {"format": "url_error", "url": url, "text": f"[Could not fetch {url}: {e}]"}


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def extract(data: bytes, content_type: str, filename: str = "") -> dict[str, Any]:
    """Extract text/structured content from file bytes based on MIME type."""
    ct = content_type.lower().split(";")[0].strip()

    if ct == "application/pdf":
        return _extract_pdf(data)

    if ct in (
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "text/csv",
        "application/csv",
    ) or filename.lower().endswith((".xlsx", ".xls", ".csv")):
        return _extract_excel(data, filename)

    if ct in (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    ) or filename.lower().endswith((".docx", ".doc")):
        return _extract_word(data)

    if ct.startswith("image/"):
        return _extract_image(data, content_type, filename)

    if ct.startswith("text/") or ct in ("application/json", "application/xml"):
        try:
            text = data.decode("utf-8", errors="replace")
            return {"format": "text", "text": text[:10000]}
        except Exception:
            return {"format": "text", "text": "[Could not decode file as text]"}

    return {"format": "binary", "text": f"[Binary file — {len(data)} bytes — cannot extract text]"}


def extract_url(url: str) -> dict[str, Any]:
    """Fetch and extract content from a URL."""
    return _extract_url(url)
